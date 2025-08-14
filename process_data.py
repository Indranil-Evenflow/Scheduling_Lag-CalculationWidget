# process_data.py

import pandas as pd
import numpy as np
import re
import os
import tempfile
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv

# --- Configuration (from your original script) ---
SCHED_LAG_COL = 'scheduling_lag'
FLAG_COL = 'evenflow_flag'
COUNT_COL = 'active_appointment_count'

HEADER_KEYWORDS = {
    SCHED_LAG_COL: ['scheduling', 'lag'],
    FLAG_COL: ['evenflow', 'flag'],
    COUNT_COL: ['count', 'appointment']
}
MIN_HEADER_CONFIDENCE_SCORE = 2
MAX_HEADER_SCAN_ROWS = 20
# --- End Configuration ---


def find_header_and_delimiter(file_path: str) -> tuple:
    """
    Scans a text-based file (CSV/TSV) to find the header row and delimiter.
    """
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        # Use csv.Sniffer to automatically detect dialect (delimiter, quotechar, etc.)
        try:
            dialect = csv.Sniffer().sniff(f.read(4096))
            delimiter = dialect.delimiter
            f.seek(0) # Reset file pointer after sniffing
        except csv.Error:
            # Sniffer can fail on some files; fall back to common delimiters
            delimiter = '\t' if '\t' in f.readline() else ','
            f.seek(0)
            
        header_row_index = -1
        best_match_score = -1

        for i, line in enumerate(f):
            if i >= MAX_HEADER_SCAN_ROWS: break
            
            line_lower = line.lower()
            score = sum(
                1 for concept, keywords in HEADER_KEYWORDS.items()
                if any(all(kw in part for kw in keywords) for part in line_lower.split(delimiter))
            )
            
            if score >= MIN_HEADER_CONFIDENCE_SCORE and score > best_match_score:
                header_row_index = i
                best_match_score = score

    if header_row_index == -1:
        raise ValueError(f"Could not dynamically locate header row in '{Path(file_path).name}'.")
        
    print(f"Header dynamically detected on row {header_row_index + 1} with delimiter '{delimiter}'.")
    return header_row_index, delimiter


def load_and_validate_data(file_path: str) -> pd.DataFrame:
    """
    Loads, cleans, and validates data, robustly handling various CSV/Excel formats.
    """
    file = Path(file_path)
    if not file.exists():
        raise FileNotFoundError(f"Error: The file '{file_path}' was not found.")

    file_suffix = file.suffix.lower()

    if file_suffix in ['.xlsx', '.xls']:
        header_idx = find_header_row(str(file)) # Re-using a simplified Excel version for now
        df = pd.read_excel(file, header=header_idx)
    elif file_suffix == '.csv':
        header_idx, delimiter = find_header_and_delimiter(str(file))
        df = pd.read_csv(file, header=0, skiprows=header_idx, sep=delimiter)
    else:
        raise ValueError(f"Unsupported file format: '{file_suffix}'.")

    if df.empty: raise ValueError("Input file is empty or data could not be read.")
    
    # --- Column Cleaning and Renaming ---
    df.columns = [str(c).lower().strip().replace(' ', '_').replace('(', '').replace(')', '') for c in df.columns]
    
    final_cols = {}
    for col_name in df.columns:
        for concept, keywords in HEADER_KEYWORDS.items():
            if all(kw in col_name for kw in keywords):
                final_cols[col_name] = concept
                break
    df.rename(columns=final_cols, inplace=True)
    
    missing_cols = [c for c in [SCHED_LAG_COL, FLAG_COL, COUNT_COL] if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Essential columns missing: {missing_cols}. Found: {list(df.columns)}")

    # --- Data Type Conversion with Robustness ---
    # NEW: Remove commas from numbers before converting
    if df[COUNT_COL].dtype == 'object':
        df[COUNT_COL] = df[COUNT_COL].str.replace(',', '', regex=False)
        
    df[SCHED_LAG_COL] = pd.to_numeric(df[SCHED_LAG_COL], errors='coerce')
    df[COUNT_COL] = pd.to_numeric(df[COUNT_COL], errors='coerce')
    
    df.dropna(subset=[SCHED_LAG_COL, FLAG_COL, COUNT_COL], inplace=True)
    df[SCHED_LAG_COL] = df[SCHED_LAG_COL].astype(int)

    return df

# Helper for Excel header finding (from your working code)
def find_header_row(file_path: str) -> int:
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    sheet = workbook.active
    best_match_score = -1
    header_row_index = -1

    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True)):
        if not any(row): continue
        row_texts = [str(cell).lower() for cell in row if cell]
        score = sum(
            1 for concept, keywords in HEADER_KEYWORDS.items()
            if any(all(kw in text for kw in keywords) for text in row_texts)
        )
        if score > best_match_score:
            best_match_score = score
            header_row_index = i

    if best_match_score < MIN_HEADER_CONFIDENCE_SCORE:
        raise ValueError(f"Could not reliably locate header in '{Path(file_path).name}'. Ensure columns with 'lag', 'flag', 'count' exist.")
    print(f"Excel header dynamically detected on row {header_row_index + 1}.")
    return header_row_index

#
# The rest of the functions (generate_analysis_tables, write_to_excel, run_full_analysis)
# remain completely unchanged as their logic is correct.
#

def generate_analysis_tables(df: pd.DataFrame) -> tuple:
    """Generates the three required analysis tables with correct logic."""
    pivot_df = df.pivot_table(
        index=SCHED_LAG_COL, columns=FLAG_COL, values=COUNT_COL,
        aggfunc='sum', fill_value=0
    )

    if pivot_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    pivot_df['Grand Total'] = pivot_df.sum(axis=1)
    grand_total_row = pivot_df.sum().to_frame('Grand Total').T
    grand_total_row.index = ['Grand Total']
    table1_pivot = pd.concat([pivot_df, grand_total_row])
    table1_pivot.index.name = 'scheduling_lag'

    calc_base = pivot_df.drop('Grand Total', axis=1)
    column_grand_totals = grand_total_row.drop('Grand Total', axis=1).iloc[0]
    running_totals_df = calc_base.cumsum()
    cumulative_pct_df = running_totals_df.div(column_grand_totals.replace(0, np.nan))
    cumulative_pct_df.columns = [f"{col} %" for col in cumulative_pct_df.columns]
    table2_analysis = pd.concat([running_totals_df, cumulative_pct_df], axis=1).reset_index()

    weekly_filter = table2_analysis[SCHED_LAG_COL] <= 7
    pct_cols = [col for col in table2_analysis.columns if '%' in col]
    table3_weekly_summary = table2_analysis.loc[weekly_filter, [SCHED_LAG_COL] + pct_cols].copy()

    return table1_pivot, table2_analysis, table3_weekly_summary


def write_to_excel(output_path: str, raw_df: pd.DataFrame, tables: tuple):
    """Writes all dataframes to a multi-sheet, formatted Excel file."""
    table1, table2, table3 = tables

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        
        if 'Analysis' in writer.book.sheetnames: del writer.book['Analysis']
        ws = writer.book.create_sheet('Analysis')
        writer.sheets['Analysis'] = ws

        ws.cell(row=1, column=1, value="Count by Schedule Lag and Flag").font = Font(bold=True, size=14)
        table1.to_excel(writer, sheet_name='Analysis', startrow=1)
        
        next_col_start = table1.shape[1] + 3
        
        ws.cell(row=1, column=next_col_start, value="Running Total & Cumulative Percentage").font = Font(bold=True, size=14)
        table2.to_excel(writer, sheet_name='Analysis', startrow=1, startcol=next_col_start - 1, index=False)
        
        next_col_start += table2.shape[1] + 2
        
        ws.cell(row=1, column=next_col_start, value="% of Total: Within 1 Week").font = Font(bold=True, size=14)
        table3.to_excel(writer, sheet_name='Analysis', startrow=1, startcol=next_col_start - 1, index=False)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        percentage_format = '0%'

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.row == 2:
                    cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal='center', vertical='center')
                if '%' in str(ws.cell(row=2, column=cell.column).value) and cell.row > 2:
                    cell.number_format = percentage_format
        
        for column_cells in ws.columns:
            length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=8)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

def run_full_analysis(uploaded_file):
    """Orchestrator function called by the Streamlit app."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(uploaded_file.name).suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    raw_data_df = load_and_validate_data(tmp_path)
    analysis_tables = generate_analysis_tables(raw_data_df)
    
    table1, table2, weekly_summary = analysis_tables
    
    output_filename = f"Scheduling_Lag_Calculation_Output_{Path(uploaded_file.name).stem}.xlsx"
    output_path = Path(tempfile.gettempdir()) / output_filename
    
    write_to_excel(str(output_path), raw_data_df, analysis_tables)
    os.unlink(tmp_path)
    
    return weekly_summary, str(output_path)